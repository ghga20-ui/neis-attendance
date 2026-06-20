"""Python API exposed to the pywebview JS frontend."""
from __future__ import annotations

import csv
import json
import logging
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

from googleapiclient.discovery import build

from subject_teacher.app_service import run_day
from subject_teacher.auth.google_oauth import (
    REAUTH_MESSAGE,
    get_credentials,
    is_reauthentication_error,
    revoke,
)
from subject_teacher.drive.schemas import (
    SCHEMA_VERSION,
    Absence,
    MarkType,
    MonthlyAttendance,
    SlotAttendance,
    Students,
    Timetable,
    TimetableSlot,
)
from subject_teacher.neis_open_api import (
    normalize_subject_name,
    query_class_timetable,
    query_subject_candidates,
    subject_matches,
)
from subject_teacher.state import (
    build_store,
    load_local_neis_api_key,
    load_local_password,
    save_local_neis_api_key,
    save_local_password,
    summarize_day,
    serialize_timetable_tsv,
    parse_timetable_tsv,
    serialize_students_tsv,
    parse_students_tsv,
)

logger = logging.getLogger(__name__)
SLOT_CACHE_TTL_SECONDS = 60
API_IO_LOCK = threading.RLock()
SERIALIZED_API_METHODS = {
    "get_drive_user",
    "get_settings",
    "save_settings",
    "get_timetable_tsv",
    "save_timetable_tsv",
    "preview_neis_public_timetable",
    "publish_neis_timetable_for_week",
    "find_neis_subject_candidates",
    "get_students_tsv",
    "save_students_tsv",
    "import_students_file",
    "get_today_slots",
    "get_mobile_snapshot",
    "get_mobile_attendance_month",
    "save_mobile_slot_attendance",
    "save_slot_attendance",
}


def _json_error(exc: Exception) -> str:
    if is_reauthentication_error(exc):
        revoke()
        return json.dumps(
            {"error": REAUTH_MESSAGE, "code": "reauth_required"},
            ensure_ascii=False,
        )
    payload = {"error": str(exc)}
    code = getattr(exc, "code", None)
    if code:
        payload["code"] = code
    return json.dumps(payload, ensure_ascii=False)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_csv_rows(path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError(f"CSV ?몄퐫?⑹쓣 ?쎌쓣 ???놁뒿?덈떎: {last_error}")

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    return [[_cell_text(cell) for cell in row] for row in csv.reader(text.splitlines(), dialect)]


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _students_from_rows(rows: Iterable[Iterable[str]]) -> list[dict[str, object]]:
    cleaned = [[_cell_text(cell) for cell in row] for row in rows]
    cleaned = [row for row in cleaned if any(row)]
    if not cleaned:
        raise ValueError("?숈깮 紐낅? ?뚯씪???쎌쓣 ?됱씠 ?놁뒿?덈떎.")

    first = [cell.lower().replace(" ", "") for cell in cleaned[0]]
    number_headers = {"번호", "번", "출석번호", "number", "no", "num"}
    name_headers = {"이름", "성명", "학생명", "name"}
    number_index = next((i for i, cell in enumerate(first) if cell in number_headers), None)
    name_index = next((i for i, cell in enumerate(first) if cell in name_headers), None)
    data_rows = cleaned
    if number_index is not None and name_index is not None:
        data_rows = cleaned[1:]
    else:
        number_index, name_index = 0, 1

    students: list[dict[str, object]] = []
    for row in data_rows:
        if len(row) <= max(number_index, name_index):
            continue
        number_text = row[number_index]
        name = row[name_index]
        if not number_text or not name:
            continue
        try:
            number = int(float(number_text))
        except ValueError:
            continue
        students.append({"n": number, "name": name})

    if not students:
        raise ValueError("?숈깮 踰덊샇? ?대쫫??李얠? 紐삵뻽?듬땲?? '踰덊샇, ?대쫫' ?댁씠 ?꾩슂?⑸땲??")
    students.sort(key=lambda item: int(item["n"]))
    return students


def _parse_student_file(path: Path, class_key: str) -> dict[str, object]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
    elif suffix == ".xlsx":
        rows = _read_xlsx_rows(path)
    else:
        raise ValueError("CSV ?먮뒗 XLSX ?뚯씪留?媛?몄삱 ???덉뒿?덈떎.")
    return {"classKey": class_key, "students": _students_from_rows(rows)}


def _empty_mobile_month(month: str) -> MonthlyAttendance:
    return MonthlyAttendance(
        schemaVersion=SCHEMA_VERSION,
        month=month,
        records={},
    )


def _dump_model_or_none(model: object | None) -> dict[str, object] | None:
    if model is None:
        return None
    return model.model_dump(by_alias=True, mode="json")


def _mobile_drive_state(sync_status: str = "idle", error: str | None = None) -> dict[str, object]:
    return {
        "authStatus": "error" if error else "connected",
        "syncStatus": "error" if error else sync_status,
        "account": None,
        "error": error,
    }


def _save_slot_attendance_record(
    date_str: str,
    slot_id: str,
    attendance: SlotAttendance,
) -> MonthlyAttendance:
    with API_IO_LOCK:
        store = build_store()
        month = date_str[:7]
        monthly = store.load_monthly(month) or _empty_mobile_month(month)
        day_records = dict(monthly.records.get(date_str, {}))
        day_records[slot_id] = attendance
        monthly.records[date_str] = day_records
        store.save_monthly(monthly)
        return monthly


def _assigned_value(lesson: object, key: str, default: object = "") -> object:
    if isinstance(lesson, dict):
        return lesson.get(key, lesson.get(key.replace("_", ""), default))
    return getattr(lesson, key, default)


def _slot_marks(attendance: SlotAttendance | None) -> dict[str, str]:
    if attendance is None:
        return {}
    return {
        str(absence.student_number): absence.mark_type.value
        for absence in attendance.absences
    }


def _neis_slot_id(grade: int, class_no: str, period: int, subject: str) -> str:
    normalized = normalize_subject_name(subject) or "subject"
    safe_class = str(class_no).replace("\t", "").replace("\r", "").replace("\n", "")
    return f"neis-{grade}-{safe_class}-{period}-{normalized}"


def _week_start(date_str: str) -> datetime:
    selected = datetime.fromisoformat(date_str)
    return selected - timedelta(days=selected.weekday())


def _neis_mode_today_slots(
    settings,
    date_str: str,
    monthly,
    *,
    school_cache: dict | None = None,
    max_workers: int = 6,
) -> list[dict[str, object]]:
    # `monthly` is supplied (and cached) by the caller; this function does no
    # Drive I/O so it is safe to call from the per-class thread pool below.
    day_records = monthly.records.get(date_str, {}) if monthly else {}
    api_key = load_local_neis_api_key()

    def rows_for_assigned(assigned) -> list[dict[str, object]]:
        grade = int(_assigned_value(assigned, "grade", 0))
        class_no = str(_assigned_value(assigned, "class_no", ""))
        subject_name = str(_assigned_value(assigned, "subject_name", ""))
        neis_label = str(_assigned_value(assigned, "neis_subject_label", "")) or subject_name
        aliases = list(_assigned_value(assigned, "subject_aliases", []) or [])
        preview = query_class_timetable(
            region=settings.region,
            school_name=settings.school_name,
            date_str=date_str,
            grade=grade,
            class_no=class_no,
            api_key=api_key,
            school_cache=school_cache,
        )
        assigned_rows: list[dict[str, object]] = []
        for lesson in preview.get("lessons", []):
            actual_subject = str(lesson.get("subject") or "")
            if not subject_matches(actual_subject, [subject_name, neis_label], aliases):
                continue
            period = int(lesson.get("period") or 0)
            slot_id = _neis_slot_id(grade, class_no, period, neis_label)
            attendance = day_records.get(slot_id)
            absence_count = len(attendance.absences) if attendance else 0
            assigned_rows.append(
                {
                    "id": slot_id,
                    "period": period,
                    "grade": grade,
                    "classNo": class_no,
                    "subject": subject_name,
                    "neisLabel": actual_subject,
                    "room": f"{grade}-{class_no}",
                    "roster": f"{grade}-{class_no}",
                    "time": f"{period}교시",
                    "note": (
                        f"결과·출석인정 {absence_count}명"
                        if absence_count
                        else "전원 출석" if attendance else ""
                    ),
                    "checked": attendance is not None,
                    "absences": absence_count,
                    "synced": attendance.synced_to_neis if attendance else False,
                    "closed": attendance.closed_on_neis if attendance else False,
                    "checkedAt": attendance.checked_at if attendance else None,
                    "marks": _slot_marks(attendance),
                    "source": "neisPublic",
                }
            )
        return assigned_rows

    assigned_lessons = list(settings.assigned_lessons)
    rows: list[dict[str, object]] = []
    if not assigned_lessons:
        return rows
    # Each assigned class is an independent NEIS request; fan them out so the
    # per-class round trips overlap instead of summing up.
    workers = max(1, min(max_workers, len(assigned_lessons)))
    if workers == 1:
        results = [rows_for_assigned(assigned) for assigned in assigned_lessons]
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            results = list(executor.map(rows_for_assigned, assigned_lessons))
    for assigned_rows in results:
        rows.extend(assigned_rows)
    rows.sort(key=lambda item: int(item["period"]))
    return rows


class Api:
    def __init__(self) -> None:
        self._window = None
        self._store_cache = None
        self._slot_cache: dict[tuple[object, ...], tuple[float, list[dict[str, object]]]] = {}
        # Shared NEIS school-code cache so repeated/parallel timetable lookups
        # resolve schoolInfo at most once per (region, school, key).
        self._school_cache: dict[tuple[str, str, str], object] = {}
        # Short-lived cache of the monthly attendance file, keyed by "YYYY-MM".
        # A Drive read costs ~1.3s, and the same month is re-read on every date
        # change and five times per week publish; this collapses those.
        self._monthly_cache: dict[str, tuple[float, object]] = {}
        # settings.json is read on every today-slots / publish call (~1.2s each);
        # cache it too. Cleared alongside the other caches on any save.
        self._settings_cache: tuple[float, object] | None = None

    def __getattribute__(self, name: str):
        attr = object.__getattribute__(self, name)
        if name in SERIALIZED_API_METHODS and callable(attr):
            def locked(*args, **kwargs):
                with API_IO_LOCK:
                    return attr(*args, **kwargs)

            return locked
        return attr

    def set_window(self, window) -> None:
        self._window = window

    def _store(self):
        if self._store_cache is None:
            self._store_cache = build_store()
        return self._store_cache

    def _clear_slot_cache(self, date_str: str | None = None) -> None:
        # Any save invalidates the settings cache (settings/timetable/roster edits).
        self._settings_cache = None
        if date_str is None:
            self._slot_cache.clear()
            self._monthly_cache.clear()
            return
        for key in list(self._slot_cache):
            if key[0] == date_str:
                self._slot_cache.pop(key, None)
        # Saving attendance changes the month file — drop its cached copy so the
        # next read reflects the write.
        self._monthly_cache.pop(date_str[:7], None)

    def _load_monthly_cached(self, store, month: str):
        now = time.monotonic()
        cached = self._monthly_cache.get(month)
        if cached and now - cached[0] < SLOT_CACHE_TTL_SECONDS:
            return cached[1]
        monthly = store.load_monthly(month)
        self._monthly_cache[month] = (now, monthly)
        return monthly

    def _load_settings_cached(self, store):
        now = time.monotonic()
        if self._settings_cache and now - self._settings_cache[0] < SLOT_CACHE_TTL_SECONDS:
            return self._settings_cache[1]
        settings = store.load_settings()
        self._settings_cache = (now, settings)
        return settings

    def _cached_neis_mode_today_slots(self, store, settings, date_str: str) -> list[dict[str, object]]:
        assigned_signature = tuple(
            (
                _assigned_value(lesson, "grade", ""),
                _assigned_value(lesson, "class_no", ""),
                _assigned_value(lesson, "subject_name", ""),
                _assigned_value(lesson, "neis_subject_label", ""),
                tuple(_assigned_value(lesson, "subject_aliases", []) or []),
            )
            for lesson in settings.assigned_lessons
        )
        cache_key = (date_str, settings.region, settings.school_name, assigned_signature)
        now = time.monotonic()
        cached = self._slot_cache.get(cache_key)
        if cached and now - cached[0] < SLOT_CACHE_TTL_SECONDS:
            return [dict(item) for item in cached[1]]
        monthly = self._load_monthly_cached(store, date_str[:7])
        rows = _neis_mode_today_slots(settings, date_str, monthly, school_cache=self._school_cache)
        self._slot_cache[cache_key] = (now, [dict(item) for item in rows])
        return rows

    # ?? helpers ??????????????????????????????????????????????????????????????

    def _push_log(self, level: str, msg: str) -> None:
        if self._window is None:
            return
        payload = json.dumps({"lv": level, "msg": msg})
        self._window.evaluate_js(
            f"window.__pushLog && window.__pushLog({payload})"
        )

    def _push_progress(self, done: int, total: int, current: str, state: str) -> None:
        if self._window is None:
            return
        payload = json.dumps({"done": done, "total": total, "current": current, "state": state})
        self._window.evaluate_js(
            f"window.__pushProgress && window.__pushProgress({payload})"
        )

    # ?? ?⑥뒪?뚮뱶 ??????????????????????????????????????????????????????????????

    def get_password(self) -> str:
        return load_local_password()

    def save_password(self, password: str) -> None:
        save_local_password(password)

    def get_neis_api_key(self) -> str:
        return load_local_neis_api_key()

    def save_neis_api_key(self, api_key: str) -> None:
        save_local_neis_api_key(api_key)

    # ?? ?ㅼ젙 ?????????????????????????????????????????????????????????????????

    def get_settings(self) -> str:
        try:
            store = self._store()
            settings = store.load_settings()
            if settings is None:
                return json.dumps({"error": "settings.json ?놁쓬"})
            return settings.model_dump_json(by_alias=True)
        except Exception as exc:
            logger.exception("get_settings failed")
            return _json_error(exc)

    def save_settings(self, payload: str) -> str:
        try:
            from subject_teacher.drive.schemas import Settings
            store = self._store()
            settings = Settings.model_validate_json(payload)
            store.save_settings(settings)
            self._clear_slot_cache()
            return json.dumps({"ok": True})
        except Exception as exc:
            logger.exception("save_settings failed")
            return _json_error(exc)

    def get_drive_user(self) -> str:
        try:
            service = build("drive", "v3", credentials=get_credentials())
            return json.dumps(
                service.about().get(fields="user").execute().get("user", {}),
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("get_drive_user failed")
            return _json_error(exc)

    # ?? ?쒓컙??????????????????????????????????????????????????????????????????

    def get_timetable_tsv(self) -> str:
        try:
            store = self._store()
            timetable = store.load_timetable()
            return serialize_timetable_tsv(timetable)
        except Exception as exc:
            logger.exception("get_timetable_tsv failed")
            return _json_error(exc)

    def save_timetable_tsv(self, tsv: str, effective_from: str) -> str:
        try:
            store = self._store()
            timetable = parse_timetable_tsv(tsv, effective_from)
            store.save_timetable(timetable)
            self._clear_slot_cache()
            return json.dumps({"ok": True})
        except Exception as exc:
            return _json_error(exc)

    def preview_neis_public_timetable(self, payload: str) -> str:
        try:
            data = json.loads(payload)
            result = query_class_timetable(
                region=str(data.get("region") or ""),
                school_name=str(data.get("schoolName") or ""),
                date_str=str(data.get("date") or ""),
                grade=int(data.get("grade") or 0),
                class_no=str(data.get("classNo") or ""),
                api_key=str(data.get("apiKey") or ""),
            )
            return json.dumps(result, ensure_ascii=False)
        except Exception as exc:
            logger.exception("preview_neis_public_timetable failed")
            return _json_error(exc)

    def publish_neis_timetable_for_week(self, date_str: str) -> str:
        try:
            store = self._store()
            settings = self._load_settings_cached(store)
            if settings is None:
                raise RuntimeError("settings.json 없음")
            if settings.timetable_mode != "neis":
                return json.dumps({"ok": True, "count": 0, "effectiveFrom": date_str})

            week_start = _week_start(date_str)
            slots: list[TimetableSlot] = []
            seen: set[tuple[str, int]] = set()

            # Iterate the five weekdays sequentially. _cached_neis_mode_today_slots
            # reads the shared Drive store (googleapiclient/httplib2), which is NOT
            # thread-safe; fanning the days out concurrently corrupts that client.
            # Speed still comes from the per-class NEIS parallelism inside each day.
            for offset in range(5):
                target_date = (week_start + timedelta(days=offset)).date().isoformat()
                for row in self._cached_neis_mode_today_slots(store, settings, target_date):
                    slot_id = str(row["id"])
                    day_of_week = offset + 1
                    dedupe_key = (slot_id, day_of_week)
                    if dedupe_key in seen:
                        continue
                    seen.add(dedupe_key)
                    slots.append(
                        TimetableSlot(
                            id=slot_id,
                            dayOfWeek=day_of_week,
                            period=int(row["period"]),
                            grade=int(row["grade"]),
                            classNo=str(row["classNo"]),
                            subjectName=str(row["subject"]),
                            neisSubjectLabel=str(row["neisLabel"]),
                        )
                    )
            slots.sort(key=lambda slot: (slot.day_of_week, slot.period, slot.grade, slot.class_no))
            timetable = Timetable(
                schemaVersion=SCHEMA_VERSION,
                effectiveFrom=week_start.date().isoformat(),
                slots=slots,
            )
            store.save_timetable(timetable)
            self._clear_slot_cache()
            return json.dumps(
                {"ok": True, "count": len(slots), "effectiveFrom": week_start.date().isoformat()},
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("publish_neis_timetable_for_week failed")
            return _json_error(exc)

    def find_neis_subject_candidates(self, payload: str) -> str:
        try:
            data = json.loads(payload)
            result = query_subject_candidates(
                region=str(data.get("region") or ""),
                school_name=str(data.get("schoolName") or ""),
                date_str=str(data.get("date") or ""),
                grade=int(data.get("grade") or 0),
                class_no=str(data.get("classNo") or ""),
                subject_name=str(data.get("subjectName") or ""),
                api_key=str(data.get("apiKey") or "") or load_local_neis_api_key(),
            )
            return json.dumps(result, ensure_ascii=False)
        except Exception as exc:
            logger.exception("find_neis_subject_candidates failed")
            return _json_error(exc)

    # ?? ?숈깮 紐낅? ?????????????????????????????????????????????????????????????

    def get_students_tsv(self) -> str:
        try:
            from subject_teacher.local_store import load_local_students

            return serialize_students_tsv(load_local_students())
        except Exception as exc:
            logger.exception("get_students_tsv failed")
            return _json_error(exc)

    def save_students_tsv(self, tsv: str) -> str:
        try:
            from subject_teacher.local_store import save_local_students

            students = parse_students_tsv(tsv)
            save_local_students(students)
            self._clear_slot_cache()
            return json.dumps({"ok": True})
        except Exception as exc:
            return _json_error(exc)

    def import_students_file(self, class_key: str) -> str:
        try:
            if not class_key:
                raise ValueError("癒쇱? ?숆툒???좏깮?섍굅??異붽???二쇱꽭??")
            if self._window is None:
                raise RuntimeError("?뚯씪 ?좏깮 李쎌쓣 ?????놁뒿?덈떎.")

            import webview

            paths = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("?숈깮 紐낅? (*.csv;*.xlsx)", "CSV (*.csv)", "Excel (*.xlsx)"),
            )
            if not paths:
                return json.dumps({"cancelled": True}, ensure_ascii=False)

            return json.dumps(_parse_student_file(Path(paths[0]), class_key), ensure_ascii=False)
        except Exception as exc:
            logger.exception("import_students_file failed")
            return _json_error(exc)

    # ?? ?ㅻ뒛 ?섏뾽 ?щ’ ????????????????????????????????????????????????????????

    def get_today_slots(self, date_str: str) -> str:
        try:
            store = self._store()
            settings = self._load_settings_cached(store)
            if settings is not None and settings.timetable_mode == "neis":
                return json.dumps(self._cached_neis_mode_today_slots(store, settings, date_str), ensure_ascii=False)
            summaries = summarize_day(store, date_str)
            monthly = self._load_monthly_cached(store, date_str[:7])
            day_records = monthly.records.get(date_str, {}) if monthly else {}
            result = []
            for s in summaries:
                attendance = day_records.get(s.slot_id)
                result.append(
                    {
                    "id": s.slot_id,
                    "period": s.period,
                    "grade": s.grade,
                    "classNo": str(s.class_no),
                    "subject": s.subject_name,
                    "neisLabel": s.neis_subject_label,
                    "room": f"{s.grade}-{s.class_no}",
                    "roster": f"{s.grade}-{s.class_no}",
                    "time": f"{s.period}교시",
                    "note": (
                        f"결과·출석인정 {s.absence_count}명"
                        if s.absence_count
                        else "전원 출석" if s.checked else ""
                    ),
                    "checked": s.checked,
                    "absences": s.absence_count,
                    "synced": s.synced_to_neis,
                    "closed": s.closed_on_neis,
                    "checkedAt": attendance.checked_at if attendance else None,
                    "marks": _slot_marks(attendance),
                    }
                )
            return json.dumps(result)
        except Exception as exc:
            logger.exception("get_today_slots failed")
            return _json_error(exc)

    def get_mobile_snapshot(self, date_str: str) -> str:
        try:
            store = self._store()
            month = date_str[:7]
            settings = store.load_settings()
            timetable = store.load_timetable() or Timetable(
                schemaVersion=SCHEMA_VERSION,
                effectiveFrom=date_str,
                slots=[],
            )
            monthly = store.load_monthly(month) or _empty_mobile_month(month)
            return json.dumps(
                {
                    "settings": _dump_model_or_none(settings),
                    "timetable": timetable.model_dump(by_alias=True, mode="json"),
                    "attendanceByDate": monthly.model_dump(by_alias=True, mode="json")["records"],
                    "queue": [],
                    "isOnline": True,
                    "syncError": None,
                    "drive": _mobile_drive_state(),
                },
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("get_mobile_snapshot failed")
            return _json_error(exc)

    def get_mobile_attendance_month(self, month: str) -> str:
        try:
            store = self._store()
            monthly = store.load_monthly(month) or _empty_mobile_month(month)
            return json.dumps(
                {
                    "attendanceByDate": monthly.model_dump(by_alias=True, mode="json")["records"],
                    "queue": [],
                    "isOnline": True,
                    "syncError": None,
                    "drive": _mobile_drive_state(),
                },
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("get_mobile_attendance_month failed")
            return _json_error(exc)

    def save_mobile_slot_attendance(self, date_str: str, slot_id: str, attendance_payload: str) -> str:
        try:
            attendance = SlotAttendance.model_validate_json(attendance_payload)
            if attendance.source != "mobile":
                attendance = attendance.model_copy(update={"source": "mobile"})
            attendance = attendance.model_copy(
                update={"synced_to_neis": False, "closed_on_neis": False}
            )
            monthly = _save_slot_attendance_record(date_str, slot_id, attendance)
            return json.dumps(
                {
                    "attendanceByDate": monthly.model_dump(by_alias=True, mode="json")["records"],
                    "queue": [],
                    "isOnline": True,
                    "syncError": None,
                    "drive": _mobile_drive_state(),
                },
                ensure_ascii=False,
            )
        except Exception as exc:
            logger.exception("save_mobile_slot_attendance failed")
            return _json_error(exc)

    def save_slot_attendance(self, date_str: str, slot_id: str, marks_payload: str) -> str:
        try:
            raw_marks = json.loads(marks_payload)
            if not isinstance(raw_marks, dict):
                raise ValueError("marks payload must be an object")

            absences: list[Absence] = []
            for number_text, mark in raw_marks.items():
                if mark in (None, "", "present"):
                    continue
                if mark not in (MarkType.ABSENT.value, MarkType.EXCUSED.value):
                    raise ValueError(f"unknown mark type: {mark!r}")
                absences.append(
                    Absence(
                        studentNumber=int(number_text),
                        markType=MarkType(mark),
                        note="",
                    )
                )
            absences.sort(key=lambda item: item.student_number)

            checked_at = datetime.now().astimezone().isoformat(timespec="seconds")
            _save_slot_attendance_record(
                date_str,
                slot_id,
                SlotAttendance(
                absences=absences,
                checkedAt=checked_at,
                source="pc",
                syncedToNeis=False,
                closedOnNeis=False,
                ),
            )
            self._clear_slot_cache(date_str)
            return json.dumps({"ok": True, "checkedAt": checked_at})
        except Exception as exc:
            logger.exception("save_slot_attendance failed")
            return _json_error(exc)

    # ?? NEIS ?ㅽ뻾 ?????????????????????????????????????????????????????????????

    def start_run(self, date_str: str, password: str, close_after: bool) -> None:
        def _worker() -> None:
            try:
                self._push_log("INFO", f"NEIS 반영 시작 — {date_str}")
                self._push_progress(0, 0, "", "running")
                results = run_day(
                    date_str,
                    password,
                    bool(close_after),
                    keep_browser_open=True,
                )
                total = len(results)
                for i, r in enumerate(results, 1):
                    if r.status == "ok":
                        self._push_log("OK", f"{r.slot_id} 반영됨")
                    elif r.status == "skipped":
                        self._push_log("INFO", f"{r.slot_id} 건너뜀 (이미 반영)")
                    else:
                        self._push_log("ERR", f"{r.slot_id} 실패: {r.error}")
                    self._push_progress(i, total, r.slot_id, "running")
                self._clear_slot_cache(date_str)
                self._push_progress(total, total, "", "done")
                self._push_log("OK", f"실행 완료 — {total}건 처리됨")
            except Exception as exc:
                logger.exception("start_run worker failed")
                self._push_log("ERR", f"실행 오류: {exc}")
                self._push_progress(0, 0, "", "error")

        threading.Thread(target=_worker, daemon=True).start()
